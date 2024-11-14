from fastapi import FastAPI, UploadFile, File, HTTPException, Request, Form
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import pandas as pd
import json
import os
from datetime import datetime
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

app = FastAPI(
    title="API Conversor JSON para Excel",
    description="API para converter arquivos JSON em Excel",
    version="1.0.0"
)

# Configurar arquivos estáticos e templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Rota principal - renderiza o template HTML
@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/convert/file/")
async def convert_file(json_file: UploadFile = File(...)):
    """
    Converte um arquivo JSON enviado para Excel
    """
    if not json_file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="Por favor, envie um arquivo JSON.")
    
    try:
        content = await json_file.read()
        json_data = json.loads(content.decode('utf-8'))
        return await convert_json_to_excel(json_data, json_file.filename)

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Arquivo JSON inválido")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

@app.post("/convert/json/")
async def convert_json(json_text: str = Form(...)):
    """
    Converte JSON enviado diretamente no body para Excel
    """
    try:
        # Tenta fazer o parse do JSON
        json_data = json.loads(json_text)
        
        # Verifica se o JSON é um dicionário ou lista
        if not isinstance(json_data, (dict, list)):
            raise HTTPException(status_code=400, detail="O JSON deve ser um objeto ou array")
            
        # Se for uma lista de objetos, converte diretamente
        if isinstance(json_data, list):
            df = pd.DataFrame(json_data)
        # Se for um objeto único, converte para lista com um item
        else:
            df = pd.DataFrame([json_data])
            
        # Cria nome único para o arquivo Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"uploads/json_data_{timestamp}.xlsx"
        
        # Cria diretório se não existir
        os.makedirs("uploads", exist_ok=True)
        
        # Salva como Excel
        df.to_excel(excel_filename, index=False)
        
        # Retorna o arquivo Excel
        return FileResponse(
            path=excel_filename,
            filename=os.path.basename(excel_filename),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="JSON inválido")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar JSON: {str(e)}")

async def convert_json_to_excel(json_data: dict, original_filename: str):
    """
    Função auxiliar para converter JSON em Excel com formatação adequada
    """
    try:
        # Converte JSON para DataFrame
        if isinstance(json_data, list):
            df = pd.DataFrame(json_data)
        else:
            df = pd.DataFrame([json_data])
        
        # Cria nome único para o arquivo Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"uploads/{os.path.splitext(original_filename)[0]}_{timestamp}.xlsx"
        
        # Cria diretório se não existir
        os.makedirs("uploads", exist_ok=True)
        
        # Salva primeiro sem índice
        df.to_excel(excel_filename, index=False)
        
        # Carrega o workbook para formatação
        wb = load_workbook(excel_filename)
        ws = wb.active
        
        # Define estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name='Arial', size=11, color="FFFFFF", bold=True)
        cell_font = Font(name='Arial', size=10)
        
        # Aplica estilos ao cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Aplica estilos às células de dados
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin', color='D4D4D4'),
                    right=Side(style='thin', color='D4D4D4'),
                    top=Side(style='thin', color='D4D4D4'),
                    bottom=Side(style='thin', color='D4D4D4')
                )
        
        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limita a largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Remove tabela existente se houver
        if len(ws.tables) > 0:
            del ws.tables[ws.tables.keys()[0]]
        
        # Cria uma nova tabela
        table_name = f"Table_{timestamp}"  # Use uma variável timestamp para nome único
        tab = Table(displayName=table_name, ref=ws.dimensions)
        
        # Aplica estilo à tabela
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        
        # Adiciona a tabela
        ws.add_table(tab)
        
        # Congela o painel
        ws.freeze_panes = 'A2'
        
        # Salva as alterações
        wb.save(excel_filename)
        
        # Retorna o arquivo Excel
        return FileResponse(
            path=excel_filename,
            filename=os.path.basename(excel_filename),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao formatar Excel: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 
